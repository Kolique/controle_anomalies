import streamlit as st
import pandas as pd
import io
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Configuration de la page Streamlit
st.set_page_config(layout="wide")
st.title("Outil de Contrôle de Données")

# #############################################################################
# --- CODE POUR L'APPLICATION 1 : RADIORELÈVE ---
# #############################################################################

def get_csv_delimiter_radio(file):
    """Détecte le délimiteur d'un fichier CSV."""
    try:
        sample = file.read(2048).decode('utf-8')
        dialect = csv.Sniffer().sniff(sample)
        file.seek(0)
        return dialect.delimiter
    except Exception:
        file.seek(0)
        return ','

def check_fp2e_details_radio(row):
    """
    Vérifie les détails FP2E et retourne les anomalies ET les corrections.
    Retour: (liste_anomalies, dict_corrections)
    """
    anomalies = []
    corrections = {}
    try:
        compteur = str(row['Numéro de compteur']).strip()
        annee_fabrication_val = str(row['Année de fabrication']).strip()
        diametre_val = row['Diametre']
        
        fp2e_regex = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'
        if not re.match(fp2e_regex, compteur):
            return [], {}

        annee_compteur = compteur[1:3]
        lettre_diam = compteur[4].upper()
        
        if annee_fabrication_val == '' or not annee_fabrication_val.isdigit() or annee_compteur != annee_fabrication_val.zfill(2):
            anomalies.append('L\'année de millésime n\'est pas conforme')
            corrections['annee'] = annee_compteur

        fp2e_map = {'A': 15, 'U': 15, 'V': 15, 'B': 20, 'C': 25, 'D': 30, 'E': 40, 'F': 50, 'G': [60, 65], 'H': 80, 'I': 100, 'J': 125, 'K': 150}
        expected_diametres = fp2e_map.get(lettre_diam, [])
        if not isinstance(expected_diametres, list): expected_diametres = [expected_diametres]

        if pd.isna(diametre_val) or diametre_val not in expected_diametres:
            anomalies.append('Le diamètre n\'est pas conforme')
            if lettre_diam == 'G':
                corrections['diametre'] = '60'
            elif expected_diametres:
                corrections['diametre'] = str(expected_diametres[0])
            
    except (TypeError, ValueError, IndexError):
        anomalies.append('Le numéro de compteur n\'est pas conforme')
    
    return anomalies, corrections

def check_data_radio(df):
    """Vérifie les données du DataFrame pour détecter les anomalies."""
    df_with_anomalies = df.copy()
    
    # Initialisation des nouvelles colonnes
    df_with_anomalies['Correction Année'] = ''
    df_with_anomalies['Correction Diamètre'] = ''
    df_with_anomalies['Correction Type Compteur'] = ''
    df_with_anomalies['Correction Marque'] = ''
    df_with_anomalies['Correction Numéro de Tête'] = ''
    df_with_anomalies['Correction Protocole Radio'] = ''

    if 'Type Compteur' not in df_with_anomalies.columns:
        st.error("La colonne 'Type Compteur' est manquante dans votre fichier.")
        st.stop()

    df_with_anomalies['Année de fabrication'] = df_with_anomalies['Année de fabrication'].astype(str).replace('nan', '', regex=False).apply(lambda x: str(int(float(x))) if x.replace('.', '', 1).isdigit() and x != '' else x).str.slice(-2).str.zfill(2)
    required_columns = ['Protocole Radio', 'Marque', 'Numéro de tête', 'Numéro de compteur', 'Latitude', 'Longitude', 'Commune', 'Année de fabrication', 'Diametre', 'Mode de relève', 'Type Compteur']
    if not all(col in df_with_anomalies.columns for col in required_columns):
        missing_columns = [col for col in required_columns if col not in df_with_anomalies.columns]; st.error(f"Colonnes requises manquantes : {', '.join(missing_columns)}"); st.stop()
    df_with_anomalies['Anomalie'] = ''
    df_with_anomalies['Anomalie Détaillée FP2E'] = ''
    for col in ['Numéro de compteur', 'Numéro de tête', 'Marque', 'Protocole Radio', 'Mode de relève', 'Type Compteur']: df_with_anomalies[col] = df_with_anomalies[col].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Latitude'] = pd.to_numeric(df_with_anomalies['Latitude'], errors='coerce'); df_with_anomalies['Longitude'] = pd.to_numeric(df_with_anomalies['Longitude'], errors='coerce')
    is_kamstrup = df_with_anomalies['Marque'].str.upper() == 'KAMSTRUP'; is_sappel = df_with_anomalies['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)']); is_itron = df_with_anomalies['Marque'].str.upper() == 'ITRON'; annee_fabrication_num = pd.to_numeric(df_with_anomalies['Année de fabrication'], errors='coerce'); df_with_anomalies['Diametre'] = pd.to_numeric(df_with_anomalies['Diametre'], errors='coerce')
    
    kamstrup_protocole_incorrect = is_kamstrup & (df_with_anomalies['Protocole Radio'].str.upper() != 'WMS')
    df_with_anomalies.loc[kamstrup_protocole_incorrect, 'Anomalie'] += 'KAMSTRUP: Protocole ≠ WMS / '
    df_with_anomalies.loc[kamstrup_protocole_incorrect, 'Correction Protocole Radio'] = 'WMS'
    
    sappel_protocole_incorrect_wms = is_sappel & (annee_fabrication_num <= 22) & (df_with_anomalies['Protocole Radio'].str.upper() != 'WMS')
    df_with_anomalies.loc[sappel_protocole_incorrect_wms, 'Anomalie'] += 'SAPPEL: Protocole ≠ WMS (année <= 22) / '
    df_with_anomalies.loc[sappel_protocole_incorrect_wms, 'Correction Protocole Radio'] = 'WMS'

    sappel_protocole_incorrect_oms = is_sappel & (annee_fabrication_num > 22) & (df_with_anomalies['Protocole Radio'].str.upper() != 'OMS')
    df_with_anomalies.loc[sappel_protocole_incorrect_oms, 'Anomalie'] += 'SAPPEL: Protocole ≠ OMS (année > 22) / '
    df_with_anomalies.loc[sappel_protocole_incorrect_oms, 'Correction Protocole Radio'] = 'OMS'
    
    df_with_anomalies.loc[df_with_anomalies['Marque'].isin(['', 'nan']), 'Anomalie'] += 'Marque manquante / '
    df_with_anomalies.loc[df_with_anomalies['Numéro de compteur'].isin(['', 'nan']), 'Anomalie'] += 'Numéro de compteur manquant / '
    df_with_anomalies.loc[df_with_anomalies['Diametre'].isnull(), 'Anomalie'] += 'Diamètre manquant / '
    df_with_anomalies.loc[df_with_anomalies['Année de fabrication'].isnull(), 'Anomalie'] += 'Année de fabrication manquante / '
    
    tete_manquante = df_with_anomalies['Numéro de tête'].isin(['', 'nan'])
    condition_tete_sappel = tete_manquante & (~is_sappel | (annee_fabrication_num >= 22)) & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE') & (~is_kamstrup)
    df_with_anomalies.loc[condition_tete_sappel, 'Anomalie'] += 'Numéro de tête manquant / '
    
    condition_tete_kamstrup = tete_manquante & is_kamstrup & (df_with_anomalies['Numéro de compteur'].str.match(r'^\d{8}$'))
    df_with_anomalies.loc[condition_tete_kamstrup, 'Anomalie'] += 'Numéro de tête manquant / '
    df_with_anomalies.loc[condition_tete_kamstrup, 'Correction Numéro de Tête'] = df_with_anomalies.loc[condition_tete_kamstrup, 'Numéro de compteur']

    df_with_anomalies.loc[df_with_anomalies['Latitude'].isnull() | df_with_anomalies['Longitude'].isnull(), 'Anomalie'] += 'Coordonnées GPS non numériques / '
    df_with_anomalies.loc[((df_with_anomalies['Latitude'] == 0) | (~df_with_anomalies['Latitude'].between(-90, 90))) | ((df_with_anomalies['Longitude'] == 0) | (~df_with_anomalies['Longitude'].between(-180, 180))), 'Anomalie'] += 'Coordonnées GPS invalides / '
    
    kamstrup_valid = is_kamstrup & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan']))
    df_with_anomalies.loc[is_kamstrup & (df_with_anomalies['Numéro de compteur'].str.len() != 8), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ 8 caractères / '
    df_with_anomalies.loc[kamstrup_valid & (df_with_anomalies['Numéro de compteur'] != df_with_anomalies['Numéro de tête']), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ Tête / '
    df_with_anomalies.loc[kamstrup_valid & (~df_with_anomalies['Numéro de compteur'].str.isdigit() | ~df_with_anomalies['Numéro de tête'].str.isdigit()), 'Anomalie'] += 'KAMSTRUP: Compteur ou Tête non numérique / '
    df_with_anomalies.loc[is_kamstrup & (~df_with_anomalies['Diametre'].between(15, 80)), 'Anomalie'] += 'KAMSTRUP: Diamètre hors plage / '
    df_with_anomalies.loc[is_sappel & (df_with_anomalies['Numéro de tête'].astype(str).str.upper().str.startswith('DME')) & (df_with_anomalies['Numéro de tête'].str.len() != 15), 'Anomalie'] += 'SAPPEL: Tête DME ≠ 15 caractères / '
    df_with_anomalies.loc[is_sappel & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE') & (~df_with_anomalies['Numéro de compteur'].str.startswith(('C', 'H'))), 'Anomalie'] += 'SAPPEL: Compteur ne commence pas par C ou H / '
    
    compteur_starts_C = df_with_anomalies['Numéro de compteur'].str.startswith('C'); marque_not_sappel_C = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (C)'
    df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (C) / '; df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Correction Marque'] = 'SAPPEL (C)'
    compteur_starts_H = df_with_anomalies['Numéro de compteur'].str.startswith('H'); marque_not_sappel_H = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (H)'
    df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (H) / '; df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Correction Marque'] = 'SAPPEL (H)'
    
    df_with_anomalies.loc[is_itron & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE') & (~df_with_anomalies['Numéro de compteur'].str.startswith(('I', 'D'))), 'Anomalie'] += 'ITRON: Compteur ne commence pas par I ou D / '
    
    is_brand_ok = is_sappel | is_itron; is_len_ok = df_with_anomalies['Numéro de compteur'].str.len() == 11
    starts_with_letter = df_with_anomalies['Numéro de compteur'].str[0].str.isalpha(); fourth_is_letter = df_with_anomalies['Numéro de compteur'].str[3].str.isalpha()
    condition_type_compteur = is_brand_ok & is_len_ok & starts_with_letter & fourth_is_letter
    rows_to_check = df_with_anomalies[condition_type_compteur].copy()
    if not rows_to_check.empty:
        sappel_rows = rows_to_check[rows_to_check['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)'])]
        if not sappel_rows.empty:
            correct_type_sappel = sappel_rows['Numéro de compteur'].str[0] + sappel_rows['Numéro de compteur'].str[3]
            incorrect_mask_sappel = sappel_rows['Type Compteur'] != correct_type_sappel
            incorrect_indices_sappel = sappel_rows[incorrect_mask_sappel].index
            if not incorrect_indices_sappel.empty:
                df_with_anomalies.loc[incorrect_indices_sappel, 'Anomalie'] += 'Incohérence Type Compteur / '; df_with_anomalies.loc[incorrect_indices_sappel, 'Correction Type Compteur'] = correct_type_sappel[incorrect_mask_sappel]
        
        itron_rows = rows_to_check[rows_to_check['Marque'].str.upper() == 'ITRON']
        if not itron_rows.empty:
            correct_type_itron = 'I' + itron_rows['Numéro de compteur'].str[3]
            incorrect_mask_itron = itron_rows['Type Compteur'] != correct_type_itron
            incorrect_indices_itron = itron_rows[incorrect_mask_itron].index
            if not incorrect_indices_itron.empty:
                df_with_anomalies.loc[incorrect_indices_itron, 'Anomalie'] += 'Incohérence Type Compteur / '; df_with_anomalies.loc[incorrect_indices_itron, 'Correction Type Compteur'] = correct_type_itron[incorrect_mask_itron]

    fp2e_regex = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'; sappel_non_manuelle_fp2e = is_sappel & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE'); manuelle_format_ok = (df_with_anomalies['Mode de relève'].str.upper() == 'MANUELLE') & (df_with_anomalies['Numéro de compteur'].str.match(fp2e_regex, na=False));
    fp2e_check_condition = sappel_non_manuelle_fp2e | manuelle_format_ok
    fp2e_results = df_with_anomalies[fp2e_check_condition].apply(check_fp2e_details_radio, axis=1)
    for index, result in fp2e_results.items():
        anomalies, corrections = result
        if anomalies: df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections: df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections: df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']

    df_with_anomalies['Anomalie'] = df_with_anomalies['Anomalie'].str.strip().str.rstrip(' /')
    anomalies_df = df_with_anomalies[(df_with_anomalies['Anomalie'] != '') | (df_with_anomalies['Correction Année'] != '') | (df_with_anomalies['Correction Diamètre'] != '') | (df_with_anomalies['Correction Type Compteur'] != '') | (df_with_anomalies['Correction Marque'] != '') | (df_with_anomalies['Correction Numéro de Tête'] != '') | (df_with_anomalies['Correction Protocole Radio'] != '')].copy()
    anomalies_df.reset_index(inplace=True); anomalies_df.rename(columns={'index': 'Index original'}, inplace=True)
    
    try:
        cols = list(anomalies_df.columns); cols.remove('Correction Année'); cols.remove('Correction Diamètre'); cols.remove('Correction Type Compteur'); cols.remove('Correction Marque'); cols.remove('Correction Numéro de Tête'); cols.remove('Correction Protocole Radio')
        pos_annee = cols.index('Année de fabrication') + 1; cols.insert(pos_annee, 'Correction Année')
        pos_diametre = cols.index('Diametre') + 1; cols.insert(pos_diametre, 'Correction Diamètre')
        pos_type = cols.index('Type Compteur') + 1; cols.insert(pos_type, 'Correction Type Compteur')
        pos_marque = cols.index('Marque') + 1; cols.insert(pos_marque, 'Correction Marque')
        pos_tete = cols.index('Numéro de tête') + 1; cols.insert(pos_tete, 'Correction Numéro de Tête')
        pos_protocole = cols.index('Protocole Radio') + 1; cols.insert(pos_protocole, 'Correction Protocole Radio')
        anomalies_df = anomalies_df[cols]
    except ValueError: pass

    return anomalies_df, anomalies_df['Anomalie'].str.split(' / ').explode().value_counts()

# #############################################################################
# --- CODE POUR L'APPLICATION 2 : TÉLÉRELÈVE ---
# #############################################################################

def get_csv_delimiter_tele(file):
    try:
        sample = file.read(2048).decode('utf-8'); dialect = csv.Sniffer().sniff(sample); file.seek(0); return dialect.delimiter
    except Exception:
        file.seek(0); return ','

def check_fp2e_details_tele(row):
    anomalies, corrections = [], {}
    try:
        compteur = str(row['Numéro de compteur']).strip(); annee_fabrication_val = str(row['Année de fabrication']).strip(); diametre_val = row['Diametre']
        fp2e_regex = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'
        if not re.match(fp2e_regex, compteur):
            anomalies.append('Format de compteur non FP2E'); return anomalies, corrections
        annee_compteur = compteur[1:3]; lettre_diam = compteur[4].upper()
        if not annee_fabrication_val or not annee_fabrication_val.isdigit() or annee_compteur != annee_fabrication_val.zfill(2):
            anomalies.append('Année millésime non conforme FP2E'); corrections['annee'] = annee_compteur
        fp2e_map = {'A': 15, 'U': 15, 'V': 15, 'B': 20, 'C': 25, 'D': 30, 'E': 40, 'F': 50, 'G': [60, 65], 'H': 80, 'I': 100, 'J': 125, 'K': 150}
        expected_diametres = fp2e_map.get(lettre_diam, []);
        if not isinstance(expected_diametres, list): expected_diametres = [expected_diametres]
        if pd.isna(diametre_val) or diametre_val not in expected_diametres:
            anomalies.append('Diamètre non conforme FP2E')
            if lettre_diam == 'G': corrections['diametre'] = '60'
            elif expected_diametres: corrections['diametre'] = str(expected_diametres[0])
    except (TypeError, ValueError, IndexError):
        anomalies.append('Erreur de format interne')
    return anomalies, corrections

def check_data_tele(df):
    df_with_anomalies = df.copy()
    df_with_anomalies['Correction Année'] = ''; df_with_anomalies['Correction Diamètre'] = ''; df_with_anomalies['Correction Type Compteur'] = ''; df_with_anomalies['Correction Marque'] = ''; df_with_anomalies['Correction Numéro de Tête'] = ''; df_with_anomalies['Correction Protocole Radio'] = ''
    
    if 'Type Compteur' not in df_with_anomalies.columns:
        st.error("La colonne 'Type Compteur' est manquante dans votre fichier."); st.stop()

    df_with_anomalies['Année de fabrication'] = df_with_anomalies['Année de fabrication'].astype(str).replace('nan', '', regex=False).apply(lambda x: str(int(float(x))) if x.replace('.', '', 1).isdigit() and x != '' else x).str.slice(-2).str.zfill(2)
    required_columns = ['Protocole Radio', 'Marque', 'Numéro de compteur', 'Numéro de tête', 'Latitude', 'Longitude', 'Année de fabrication', 'Diametre', 'Traité', 'Mode de relève', 'Type Compteur']
    if not all(col in df_with_anomalies.columns for col in required_columns):
        missing = [col for col in required_columns if col not in df_with_anomalies.columns]; st.error(f"Colonnes requises manquantes : {', '.join(missing)}"); st.stop()
    df_with_anomalies['Anomalie'] = ''
    for col in ['Numéro de compteur', 'Numéro de tête', 'Marque', 'Protocole Radio', 'Traité', 'Mode de relève', 'Type Compteur']: df_with_anomalies[col] = df_with_anomalies[col].astype(str).replace('nan', '', regex=False)
    df_with_anomalies['Latitude'] = pd.to_numeric(df_with_anomalies['Latitude'], errors='coerce'); df_with_anomalies['Longitude'] = pd.to_numeric(df_with_anomalies['Longitude'], errors='coerce'); df_with_anomalies['Diametre'] = pd.to_numeric(df_with_anomalies['Diametre'], errors='coerce')
    is_kamstrup = df_with_anomalies['Marque'].str.upper() == 'KAMSTRUP'; is_sappel = df_with_anomalies['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)', 'SAPPEL(C)']); is_itron = df_with_anomalies['Marque'].str.upper() == 'ITRON'; is_kaifa = df_with_anomalies['Marque'].str.upper() == 'KAIFA'; is_mode_manuelle = df_with_anomalies['Mode de relève'].str.upper() == 'MANUELLE'; annee_fabrication_num = pd.to_numeric(df_with_anomalies['Année de fabrication'], errors='coerce')
    
    traite_lra_condition = df_with_anomalies['Traité'].str.startswith(('903', '863'), na=False)
    protocole_incorrect_lra = (~is_mode_manuelle) & traite_lra_condition & (df_with_anomalies['Protocole Radio'].str.upper() != 'LRA')
    df_with_anomalies.loc[protocole_incorrect_lra, 'Anomalie'] += 'Protocole incorrect (devrait être LRA) / '
    df_with_anomalies.loc[protocole_incorrect_lra, 'Correction Protocole Radio'] = 'LRA'
    protocole_incorrect_sgx = (~is_mode_manuelle) & (~traite_lra_condition) & (df_with_anomalies['Protocole Radio'].str.upper() != 'SGX')
    df_with_anomalies.loc[protocole_incorrect_sgx, 'Anomalie'] += 'Protocole incorrect (devrait être SGX) / '
    df_with_anomalies.loc[protocole_incorrect_sgx, 'Correction Protocole Radio'] = 'SGX'

    df_with_anomalies.loc[df_with_anomalies['Marque'].isin(['', 'nan']), 'Anomalie'] += 'Marque manquante / '; df_with_anomalies.loc[df_with_anomalies['Numéro de compteur'].isin(['', 'nan']), 'Anomalie'] += 'Numéro de compteur manquant / '; df_with_anomalies.loc[df_with_anomalies['Diametre'].isnull(), 'Anomalie'] += 'Diamètre manquant / '; df_with_anomalies.loc[annee_fabrication_num.isnull(), 'Anomalie'] += 'Année de fabrication manquante / '
    df_with_anomalies.loc[df_with_anomalies['Numéro de tête'].isin(['', 'nan']) & (~is_kamstrup) & (~is_kaifa) & (~is_mode_manuelle), 'Anomalie'] += 'Numéro de tête manquant / '
    df_with_anomalies.loc[df_with_anomalies['Latitude'].isnull() | df_with_anomalies['Longitude'].isnull(), 'Anomalie'] += 'Coordonnées GPS non numériques / '; df_with_anomalies.loc[((df_with_anomalies['Latitude'] == 0) | (~df_with_anomalies['Latitude'].between(-90, 90))) | ((df_with_anomalies['Longitude'] == 0) | (~df_with_anomalies['Longitude'].between(-180, 180))), 'Anomalie'] += 'Coordonnées GPS invalides / '
    kamstrup_valid = is_kamstrup & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan'])); df_with_anomalies.loc[is_kamstrup & (df_with_anomalies['Numéro de compteur'].str.len() != 8), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ 8 caractères / '; df_with_anomalies.loc[kamstrup_valid & (df_with_anomalies['Numéro de compteur'] != df_with_anomalies['Numéro de tête']), 'Anomalie'] += 'KAMSTRUP: Compteur ≠ Tête / '; df_with_anomalies.loc[kamstrup_valid & (~df_with_anomalies['Numéro de compteur'].str.isdigit() | ~df_with_anomalies['Numéro de tête'].str.isdigit()), 'Anomalie'] += 'KAMSTRUP: Compteur ou Tête non numérique / '; df_with_anomalies.loc[is_kamstrup & (~df_with_anomalies['Diametre'].between(15, 80)), 'Anomalie'] += 'KAMSTRUP: Diamètre hors de la plage [15, 80] / '
    df_with_anomalies.loc[is_sappel & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan'])) & (df_with_anomalies['Numéro de tête'].str.len() != 16), 'Anomalie'] += 'SAPPEL: Tête ≠ 16 caractères / ';
    
    compteur_starts_C = df_with_anomalies['Numéro de compteur'].str.startswith('C'); marque_not_sappel_C = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (C)'
    df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (C) / '; df_with_anomalies.loc[is_sappel & compteur_starts_C & marque_not_sappel_C, 'Correction Marque'] = 'SAPPEL (C)'
    compteur_starts_H = df_with_anomalies['Numéro de compteur'].str.startswith('H'); marque_not_sappel_H = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (H)'
    df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (H) / '; df_with_anomalies.loc[is_sappel & compteur_starts_H & marque_not_sappel_H, 'Correction Marque'] = 'SAPPEL (H)'

    df_with_anomalies.loc[is_itron & (~df_with_anomalies['Numéro de tête'].isin(['', 'nan'])) & (df_with_anomalies['Numéro de tête'].str.len() != 8), 'Anomalie'] += 'ITRON: Tête ≠ 8 caractères / '
    
    is_brand_ok = is_sappel | is_itron; is_len_ok = df_with_anomalies['Numéro de compteur'].str.len() == 11
    starts_with_letter = df_with_anomalies['Numéro de compteur'].str[0].str.isalpha(); fourth_is_letter = df_with_anomalies['Numéro de compteur'].str[3].str.isalpha()
    condition_type_compteur = is_brand_ok & is_len_ok & starts_with_letter & fourth_is_letter
    rows_to_check = df_with_anomalies[condition_type_compteur].copy()
    if not rows_to_check.empty:
        sappel_rows = rows_to_check[rows_to_check['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)', 'SAPPEL(C)'])]
        if not sappel_rows.empty:
            correct_type_sappel = sappel_rows['Numéro de compteur'].str[0] + sappel_rows['Numéro de compteur'].str[3]
            incorrect_mask_sappel = sappel_rows['Type Compteur'] != correct_type_sappel
            incorrect_indices_sappel = sappel_rows[incorrect_mask_sappel].index
            if not incorrect_indices_sappel.empty:
                df_with_anomalies.loc[incorrect_indices_sappel, 'Anomalie'] += 'Incohérence Type Compteur / '; df_with_anomalies.loc[incorrect_indices_sappel, 'Correction Type Compteur'] = correct_type_sappel[incorrect_mask_sappel]
        
        itron_rows = rows_to_check[rows_to_check['Marque'].str.upper() == 'ITRON']
        if not itron_rows.empty:
            correct_type_itron = 'I' + itron_rows['Numéro de compteur'].str[3]
            incorrect_mask_itron = itron_rows['Type Compteur'] != correct_type_itron
            incorrect_indices_itron = itron_rows[incorrect_mask_itron].index
            if not incorrect_indices_itron.empty:
                df_with_anomalies.loc[incorrect_indices_itron, 'Anomalie'] += 'Incohérence Type Compteur / '; df_with_anomalies.loc[incorrect_indices_itron, 'Correction Type Compteur'] = correct_type_itron[incorrect_mask_itron]
    
    fp2e_regex = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'; sappel_itron_non_manuelle = (is_sappel | is_itron) & (df_with_anomalies['Mode de relève'].str.upper() != 'MANUELLE'); manuelle_format_ok = (df_with_anomalies['Mode de relève'].str.upper() == 'MANUELLE') & (df_with_anomalies['Numéro de compteur'].str.match(fp2e_regex, na=False)); fp2e_check_condition = sappel_itron_non_manuelle | manuelle_format_ok; fp2e_results = df_with_anomalies[fp2e_check_condition].apply(check_fp2e_details_tele, axis=1)
    for index, result in fp2e_results.items():
        anomalies, corrections = result
        if anomalies: df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections: df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections: df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']
    is_fp2e_compliant = df_with_anomalies['Numéro de compteur'].str.match(fp2e_regex, na=False)
    df_with_anomalies.loc[is_mode_manuelle & is_itron & is_fp2e_compliant & (~df_with_anomalies['Numéro de compteur'].str.lower().str.startswith(('i', 'd'), na=False)), 'Anomalie'] += 'ITRON manuel: doit commencer par "I" ou "D" / '
    df_with_anomalies.loc[is_mode_manuelle & is_sappel & is_fp2e_compliant & (~df_with_anomalies['Numéro de compteur'].str.lower().str.startswith(('c', 'h'), na=False)), 'Anomalie'] += 'SAPPEL manuel: doit commencer par "C" ou "H" / '
    df_with_anomalies['Anomalie'] = df_with_anomalies['Anomalie'].str.strip().str.rstrip(' /'); anomalies_df = df_with_anomalies[(df_with_anomalies['Anomalie'] != '') | (df_with_anomalies['Correction Année'] != '') | (df_with_anomalies['Correction Diamètre'] != '')| (df_with_anomalies['Correction Type Compteur'] != '') | (df_with_anomalies['Correction Marque'] != '') | (df_with_anomalies['Correction Numéro de Tête'] != '') | (df_with_anomalies['Correction Protocole Radio'] != '')].copy(); anomalies_df.reset_index(inplace=True); anomalies_df.rename(columns={'index': 'Index original'}, inplace=True)
    
    try:
        cols = list(anomalies_df.columns); cols.remove('Correction Année'); cols.remove('Correction Diamètre'); cols.remove('Correction Type Compteur'); cols.remove('Correction Marque'); cols.remove('Correction Numéro de Tête'); cols.remove('Correction Protocole Radio')
        pos_annee = cols.index('Année de fabrication') + 1; cols.insert(pos_annee, 'Correction Année')
        pos_diametre = cols.index('Diametre') + 1; cols.insert(pos_diametre, 'Correction Diamètre')
        pos_type = cols.index('Type Compteur') + 1; cols.insert(pos_type, 'Correction Type Compteur')
        pos_marque = cols.index('Marque') + 1; cols.insert(pos_marque, 'Correction Marque')
        pos_tete = cols.index('Numéro de tête') + 1; cols.insert(pos_tete, 'Correction Numéro de Tête')
        pos_protocole = cols.index('Protocole Radio') + 1; cols.insert(pos_protocole, 'Correction Protocole Radio')
        anomalies_df = anomalies_df[cols]
    except ValueError: pass

    return anomalies_df, anomalies_df['Anomalie'].str.split(' / ').explode().value_counts()

def afficher_resume_anomalies_tele(anomaly_counter):
    if not anomaly_counter.empty:
        st.subheader("Récapitulatif des anomalies"); st.dataframe(pd.DataFrame(anomaly_counter).reset_index().rename(columns={"index": "Type d'anomalie", 0: "Nombre de cas"}))

def create_summary_with_corrections(anomalies_df, anomaly_counter, tab_type="radio"):
    summary_data = []
    correction_map = {}

    if tab_type == "radio":
        correction_map = {
            'L\'année de millésime n\'est pas conforme': 'Correction Année', 'Le diamètre n\'est pas conforme': 'Correction Diamètre',
            'Incohérence Type Compteur': 'Correction Type Compteur', 'SAPPEL: Incohérence Marque/Compteur (C)': 'Correction Marque',
            'SAPPEL: Incohérence Marque/Compteur (H)': 'Correction Marque', 'Numéro de tête manquant': 'Correction Numéro de Tête',
            'KAMSTRUP: Protocole ≠ WMS': 'Correction Protocole Radio', 'SAPPEL: Protocole ≠ OMS (année > 22)': 'Correction Protocole Radio',
            'SAPPEL: Protocole ≠ WMS (année <= 22)': 'Correction Protocole Radio'
        }
    elif tab_type == "tele":
        correction_map = {
            'Année millésime non conforme FP2E': 'Correction Année', 'Diamètre non conforme FP2E': 'Correction Diamètre',
            'Incohérence Type Compteur': 'Correction Type Compteur', 'SAPPEL: Incohérence Marque/Compteur (C)': 'Correction Marque',
            'SAPPEL: Incohérence Marque/Compteur (H)': 'Correction Marque',
            'Protocole incorrect (devrait être LRA)': 'Correction Protocole Radio',
            'Protocole incorrect (devrait être SGX)': 'Correction Protocole Radio'
        }
    elif tab_type == "manuelle":
        correction_map = {
            'L\'année de millésime n\'est pas conforme': 'Correction Année',
            'Le diamètre n\'est pas conforme': 'Correction Diamètre',
            'SAPPEL: Incohérence Marque/Compteur (C)': 'Correction Marque',
            'SAPPEL: Incohérence Marque/Compteur (H)': 'Correction Marque',
            'ITRON: Incohérence Marque/Compteur': 'Correction Marque'
        }


    for anomaly_type, count in anomaly_counter.items():
        correction_col = correction_map.get(anomaly_type)
        corrections_count = 0
        if correction_col:
            corrections_count = anomalies_df[anomalies_df['Anomalie'].str.contains(re.escape(anomaly_type), na=False) & (anomalies_df[correction_col] != '')].shape[0]
        summary_data.append([anomaly_type, count, corrections_count])

    summary_df = pd.DataFrame(summary_data, columns=["Type d'anomalie", "Nombre de cas", "Corrections Proposées"])
    return summary_df
    
def check_data_manuelle(df):
    """Vérifie les données du DataFrame pour l'onglet Manuelle."""
    required_cols = ['Latitude', 'Longitude', 'Numéro de compteur', 'Marque', 'Année de fabrication', 'Diametre']
    if not all(col in df.columns for col in required_cols):
        missing = [col for col in required_cols if col not in df.columns]; st.error(f"Colonnes requises manquantes : {', '.join(missing)}"); st.stop()
        
    df_with_anomalies = df.copy()
    df_with_anomalies['Anomalie'] = ''
    df_with_anomalies['Correction Année'] = ''
    df_with_anomalies['Correction Diamètre'] = ''
    df_with_anomalies['Correction Marque'] = ''

    df_with_anomalies['Année de fabrication'] = df_with_anomalies['Année de fabrication'].astype(str).replace('nan', '', regex=False).apply(lambda x: str(int(float(x))) if x.replace('.', '', 1).isdigit() and x != '' else x).str.slice(-2).str.zfill(2)
    df_with_anomalies['Latitude'] = pd.to_numeric(df_with_anomalies['Latitude'], errors='coerce')
    df_with_anomalies['Longitude'] = pd.to_numeric(df_with_anomalies['Longitude'], errors='coerce')
    df_with_anomalies['Diametre'] = pd.to_numeric(df_with_anomalies['Diametre'], errors='coerce')

    df_with_anomalies.loc[df_with_anomalies['Latitude'].isnull() | df_with_anomalies['Longitude'].isnull(), 'Anomalie'] += 'Coordonnées GPS non numériques / '
    coord_invalid = ((df_with_anomalies['Latitude'] == 0) | (~df_with_anomalies['Latitude'].between(-90, 90))) | ((df_with_anomalies['Longitude'] == 0) | (~df_with_anomalies['Longitude'].between(-180, 180)))
    df_with_anomalies.loc[coord_invalid, 'Anomalie'] += 'Coordonnées GPS invalides / '

    is_sappel = df_with_anomalies['Marque'].str.upper().isin(['SAPPEL (C)', 'SAPPEL (H)']); is_itron = df_with_anomalies['Marque'].str.upper() == 'ITRON'
    fp2e_regex = r'^[A-Z]\d{2}[A-Z]{2}\d{6}$'
    has_fp2e_format = df_with_anomalies['Numéro de compteur'].str.match(fp2e_regex, na=False)
    
    df_with_anomalies.loc[(is_sappel | is_itron) & (~has_fp2e_format), 'Anomalie'] += 'Compteur non-FP2E pour SAPPEL/ITRON / '

    compteur_starts_C = df_with_anomalies['Numéro de compteur'].str.startswith('C'); marque_not_sappel_C = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (C)'
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_C & marque_not_sappel_C, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (C) / '
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_C & marque_not_sappel_C, 'Correction Marque'] = 'SAPPEL (C)'
    
    compteur_starts_H = df_with_anomalies['Numéro de compteur'].str.startswith('H'); marque_not_sappel_H = df_with_anomalies['Marque'].str.upper() != 'SAPPEL (H)'
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_H & marque_not_sappel_H, 'Anomalie'] += 'SAPPEL: Incohérence Marque/Compteur (H) / '
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_H & marque_not_sappel_H, 'Correction Marque'] = 'SAPPEL (H)'

    compteur_starts_ID = df_with_anomalies['Numéro de compteur'].str.startswith(('I', 'D')); marque_not_itron = df_with_anomalies['Marque'].str.upper() != 'ITRON'
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_ID & marque_not_itron, 'Anomalie'] += 'ITRON: Incohérence Marque/Compteur / '
    df_with_anomalies.loc[has_fp2e_format & compteur_starts_ID & marque_not_itron, 'Correction Marque'] = 'ITRON'

    fp2e_results = df_with_anomalies[has_fp2e_format].apply(check_fp2e_details_radio, axis=1)
    for index, result in fp2e_results.items():
        anomalies, corrections = result
        if anomalies: df_with_anomalies.loc[index, 'Anomalie'] += ' / '.join(anomalies) + ' / '
        if 'annee' in corrections: df_with_anomalies.loc[index, 'Correction Année'] = corrections['annee']
        if 'diametre' in corrections: df_with_anomalies.loc[index, 'Correction Diamètre'] = corrections['diametre']
        
    df_with_anomalies['Anomalie'] = df_with_anomalies['Anomalie'].str.strip().str.rstrip(' /')
    anomalies_df = df_with_anomalies[(df_with_anomalies['Anomalie'] != '') | (df_with_anomalies['Correction Année'] != '') | (df_with_anomalies['Correction Diamètre'] != '') | (df_with_anomalies['Correction Marque'] != '')].copy()
    
    if not anomalies_df.empty:
        anomalies_df.reset_index(inplace=True); anomalies_df.rename(columns={'index': 'Index original'}, inplace=True)
        try:
            cols = list(anomalies_df.columns); cols.remove('Correction Année'); cols.remove('Correction Diamètre'); cols.remove('Correction Marque')
            pos_annee = cols.index('Année de fabrication') + 1; cols.insert(pos_annee, 'Correction Année')
            pos_diametre = cols.index('Diametre') + 1; cols.insert(pos_diametre, 'Correction Diamètre')
            pos_marque = cols.index('Marque') + 1; cols.insert(pos_marque, 'Correction Marque')
            anomalies_df = anomalies_df[cols]
        except ValueError: pass
    
    anomaly_counter = anomalies_df['Anomalie'].str.split(' / ').explode().value_counts()
    return anomalies_df, anomaly_counter

# #############################################################################
# --- CRÉATION DES ONGLETS ET INTERFACE UTILISATEUR ---
# #############################################################################

tab1, tab2, tab3 = st.tabs(["📊 Contrôle Radiorelève", "📡 Contrôle Télérelève", "✍️ Controle manuelle"])

# --- ONGLET 1 : RADIORELÈVE (INTERFACE UTILISATEUR) ---
with tab1:
    st.header("Contrôle des données de Radiorelève")
    st.markdown("Veuillez téléverser votre fichier pour lancer les contrôles.")
    uploaded_file_radio = st.file_uploader("Choisissez un fichier (Radiorelève)", type=['csv', 'xlsx'], key="uploader_radio")
    if uploaded_file_radio:
        st.success("Fichier chargé avec succès !");
        try:
            file_extension = uploaded_file_radio.name.split('.')[-1]; dtype_mapping = {'Numéro de branchement': str, 'Abonnement': str}
            if file_extension == 'csv': df = pd.read_csv(uploaded_file_radio, sep=get_csv_delimiter_radio(uploaded_file_radio), dtype=dtype_mapping)
            elif file_extension == 'xlsx': df = pd.read_excel(uploaded_file_radio, dtype=dtype_mapping)
            st.subheader("Aperçu des 5 premières lignes"); st.dataframe(df.head())
            if st.button("Lancer les contrôles (Radiorelève)", key="button_radio"):
                with st.spinner("Contrôles en cours..."): anomalies_df, anomaly_counter = check_data_radio(df)
                if not anomalies_df.empty:
                    st.error(f"Anomalies et/ou corrections détectées : {len(anomalies_df)} lignes concernées."); anomalies_df_display = anomalies_df.drop(columns=['Anomalie Détaillée FP2E'], errors='ignore'); st.dataframe(anomalies_df_display); 
                    summary_df = create_summary_with_corrections(anomalies_df, anomaly_counter, tab_type="radio")
                    st.subheader("Récapitulatif des anomalies"); st.dataframe(summary_df)
                    anomaly_columns_map = {"KAMSTRUP: Protocole ≠ WMS": ['Protocole Radio'], "SAPPEL: Protocole ≠ OMS (année > 22)": ['Protocole Radio'], "SAPPEL: Protocole ≠ WMS (année <= 22)": ['Protocole Radio'], "Marque manquante": ['Marque'], "Numéro de compteur manquant": ['Numéro de compteur'], "Numéro de tête manquant": ['Numéro de tête'], "Coordonnées GPS non numériques": ['Latitude', 'Longitude'], "Coordonnées GPS invalides": ['Latitude', 'Longitude'], "Diamètre manquant": ['Diametre'], "Année de fabrication manquante": ['Année de fabrication'], "KAMSTRUP: Compteur ≠ 8 caractères": ['Numéro de compteur'], "KAMSTRUP: Compteur ≠ Tête": ['Numéro de compteur', 'Numéro de tête'], "KAMSTRUP: Compteur ou Tête non numérique": ['Numéro de compteur', 'Numéro de tête'], "KAMSTRUP: Diamètre hors plage": ['Diametre'], "SAPPEL: Tête DME ≠ 15 caractères": ['Numéro de tête'], "SAPPEL: Compteur ne commence pas par C ou H": ['Numéro de compteur'], "SAPPEL: Incohérence Marque/Compteur (C)": ['Marque'], "SAPPEL: Incohérence Marque/Compteur (H)": ['Marque'], "ITRON: Compteur ne commence pas par I ou D": ['Numéro de compteur'], "Le numéro de compteur n'est pas conforme": ['Numéro de compteur'], "Le diamètre n'est pas conforme": ['Diametre'], "L'année de millésime n'est pas conforme": ['Année de fabrication'], "Incohérence Type Compteur": ['Type Compteur']}
                    if file_extension == 'csv': st.download_button(label="📥 Télécharger le rapport en CSV", data=anomalies_df_display.to_csv(index=False, sep=get_csv_delimiter_radio(uploaded_file_radio)).encode('utf-8'), file_name='anomalies_radioreleve.csv', mime='text/csv')
                    elif file_extension == 'xlsx':
                        excel_buffer = io.BytesIO(); wb = Workbook();
                        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
                        ws_summary = wb.create_sheet(title="Récapitulatif", index=0); ws_all_anomalies = wb.create_sheet(title="Toutes_Anomalies", index=1)
                        for r in dataframe_to_rows(anomalies_df_display, index=False, header=True): ws_all_anomalies.append(r)
                        header_font = Font(bold=True); red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        for cell in ws_all_anomalies[1]: cell.font = header_font
                        for row_num_all, df_row in enumerate(anomalies_df.iterrows()):
                            for anomaly in str(df_row[1]['Anomalie']).split(' / '):
                                if anomaly.strip() in anomaly_columns_map:
                                    for col_name in anomaly_columns_map[anomaly.strip()]:
                                        try: ws_all_anomalies.cell(row=row_num_all + 2, column=list(anomalies_df_display.columns).index(col_name) + 1).fill = red_fill
                                        except ValueError: pass
                        for col in ws_all_anomalies.columns: ws_all_anomalies.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(cell.value)) for cell in col if cell.value) + 2
                        ws_summary['A1'] = "Récapitulatif des anomalies"; ws_summary['A1'].font = Font(bold=True, size=16); ws_summary.append([]); 
                        for r_idx, row_data in enumerate(dataframe_to_rows(summary_df, index=False, header=True)):
                            ws_summary.append(row_data)
                        for cell in ws_summary[3]: cell.font = header_font
                        created_sheet_names = {"Récapitulatif", "Toutes_Anomalies"}
                        link_row = ws_summary.max_row + 2; ws_summary.cell(row=link_row, column=1, value="Toutes les anomalies").hyperlink = f"#Toutes_Anomalies!A1"; ws_summary.cell(row=link_row, column=1).font = Font(underline="single", color="0563C1"); ws_summary.cell(row=link_row, column=2, value=len(anomalies_df))
                        for idx, (anomaly_type, count, corrections) in enumerate(summary_df.values):
                            current_row_num = 4 + idx
                            sheet_name = re.sub(r'[\\/?*\[\]:()\'"<>|]', '', anomaly_type[:28]).replace(' ', '_').strip(); original_sheet_name = sheet_name; s_counter = 1
                            while sheet_name in created_sheet_names: sheet_name = f"{original_sheet_name[:28]}_{s_counter}"; s_counter += 1
                            created_sheet_names.add(sheet_name)
                            summary_cell = ws_summary.cell(row=current_row_num, column=1)
                            summary_cell.hyperlink = f"#'{sheet_name}'!A1"; summary_cell.font = Font(underline="single", color="0563C1")
                            ws_detail = wb.create_sheet(title=sheet_name); filtered_df = anomalies_df[anomalies_df['Anomalie'].str.contains(re.escape(anomaly_type), regex=True)]; filtered_df_display = filtered_df.drop(columns=['Anomalie Détaillée FP2E'], errors='ignore')
                            for r in dataframe_to_rows(filtered_df_display, index=False, header=True): ws_detail.append(r)
                            for cell in ws_detail[1]: cell.font = header_font
                            for row_num_detail, df_row_detail in enumerate(filtered_df.iterrows()):
                                for anomaly in str(df_row_detail[1]['Anomalie']).split(' / '):
                                    if anomaly.strip() in anomaly_columns_map:
                                        for col_name in anomaly_columns_map[anomaly.strip()]:
                                            try: ws_detail.cell(row=row_num_detail + 2, column=list(filtered_df_display.columns).index(col_name) + 1).fill = red_fill
                                            except ValueError: pass
                            for col in ws_detail.columns: ws_detail.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(cell.value)) for cell in col if cell.value) + 2
                        wb.save(excel_buffer); st.download_button(label="📥 Télécharger le rapport (.xlsx)", data=excel_buffer, file_name='anomalies_radioreleve.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                else: st.success("✅ Aucune anomalie détectée.")
        except Exception as e: st.error(f"Une erreur est survenue : {e}")

# --- ONGLET 2 : TÉLÉRELÈVE (INTERFACE UTILISATEUR) ---
with tab2:
    st.header("Contrôle des données de Télérelève")
    st.markdown("Veuillez téléverser votre fichier pour lancer les contrôles.")
    uploaded_file_tele = st.file_uploader("Choisissez un fichier (Télérelève)", type=['csv', 'xlsx'], key="uploader_tele")
    if uploaded_file_tele:
        st.success("Fichier chargé avec succès !");
        try:
            file_extension = uploaded_file_tele.name.split('.')[-1]; dtype_mapping = {'Numéro de branchement': str, 'Abonnement': str}
            if file_extension == 'csv': df = pd.read_csv(uploaded_file_tele, sep=get_csv_delimiter_tele(uploaded_file_tele), dtype=dtype_mapping)
            elif file_extension == 'xlsx': df = pd.read_excel(uploaded_file_tele, dtype=dtype_mapping)
            st.subheader("Aperçu des 5 premières lignes"); st.dataframe(df.head())
            if st.button("Lancer les contrôles (Télérelève)", key="button_tele"):
                with st.spinner("Contrôles en cours..."): anomalies_df, anomaly_counter = check_data_tele(df)
                if not anomalies_df.empty:
                    st.error(f"Anomalies et/ou corrections détectées : {len(anomalies_df)} lignes concernées."); anomalies_df_display = anomalies_df.drop(columns=['Anomalie Détaillée FP2E'], errors='ignore'); st.dataframe(anomalies_df_display); 
                    summary_df = create_summary_with_corrections(anomalies_df, anomaly_counter, tab_type="tele")
                    st.subheader("Récapitulatif des anomalies"); st.dataframe(summary_df)
                    anomaly_columns_map = {"Protocole incorrect (devrait être LRA)": ['Protocole Radio'], "Protocole incorrect (devrait être SGX)": ['Protocole Radio'], "Marque manquante": ['Marque'],"Numéro de compteur manquant": ['Numéro de compteur'],"Numéro de tête manquant": ['Numéro de tête'],"Coordonnées GPS non numériques": ['Latitude', 'Longitude'],"Coordonnées GPS invalides": ['Latitude', 'Longitude'],"Diamètre manquant": ['Diametre'],"Année de fabrication manquante": ['Année de fabrication'],"KAMSTRUP: Compteur ≠ 8 caractères": ['Numéro de compteur'],"KAMSTRUP: Compteur ≠ Tête": ['Numéro de compteur', 'Numéro de tête'],"KAMSTRUP: Compteur ou Tête non numérique": ['Numéro de compteur', 'Numéro de tête'],"KAMSTRUP: Diamètre hors de la plage [15, 80]": ['Diametre'],"SAPPEL: Tête ≠ 16 caractères": ['Numéro de tête'],"SAPPEL: Incohérence Marque/Compteur (C)": ['Marque'],"SAPPEL: Incohérence Marque/Compteur (H)": ['Marque'],"ITRON: Tête ≠ 8 caractères": ['Numéro de tête'],"ITRON manuel: doit commencer par \"I\" ou \"D\"": ['Numéro de compteur'],"SAPPEL manuel: doit commencer par \"C\" ou \"H\"": ['Numéro de compteur'],"Format de compteur non FP2E": ['Numéro de compteur'],"Année millésime non conforme FP2E": ['Année de fabrication'],"Diamètre non conforme FP2E": ['Diametre'], "Incohérence Type Compteur": ['Type Compteur']}
                    if file_extension == 'csv': st.download_button(label="📥 Télécharger le rapport en CSV", data=anomalies_df_display.to_csv(index=False, sep=get_csv_delimiter_tele(uploaded_file_tele)).encode('utf-8'), file_name='anomalies_telerelève.csv', mime='text/csv')
                    elif file_extension == 'xlsx':
                        excel_buffer = io.BytesIO(); wb = Workbook();
                        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
                        ws_summary = wb.create_sheet(title="Récapitulatif", index=0); ws_all_anomalies = wb.create_sheet(title="Toutes_Anomalies", index=1)
                        for r in dataframe_to_rows(anomalies_df_display, index=False, header=True): ws_all_anomalies.append(r)
                        header_font = Font(bold=True); red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        for cell in ws_all_anomalies[1]: cell.font = header_font
                        for row_num_all, df_row in enumerate(anomalies_df.iterrows()):
                            for anomaly in str(df_row[1]['Anomalie']).split(' / '):
                                if anomaly.strip() in anomaly_columns_map:
                                    for col_name in anomaly_columns_map[anomaly.strip()]:
                                        try: ws_all_anomalies.cell(row=row_num_all + 2, column=list(anomalies_df_display.columns).index(col_name) + 1).fill = red_fill
                                        except ValueError: pass
                        for col in ws_all_anomalies.columns: ws_all_anomalies.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(cell.value)) for cell in col if cell.value) + 2
                        ws_summary['A1'] = "Récapitulatif des anomalies"; ws_summary['A1'].font = Font(bold=True, size=16); ws_summary.append([]); 
                        for r in dataframe_to_rows(summary_df, index=False, header=True): ws_summary.append(r)
                        for cell in ws_summary[3]: cell.font = header_font
                        created_sheet_names = {"Récapitulatif", "Toutes_Anomalies"}
                        link_row = ws_summary.max_row + 2; ws_summary.cell(row=link_row, column=1, value="Toutes les anomalies").hyperlink = f"#'Toutes_Anomalies'!A1"; ws_summary.cell(row=link_row, column=1).font = Font(underline="single", color="0563C1"); ws_summary.cell(row=link_row, column=2, value=len(anomalies_df))
                        for idx, (anomaly_type, count, corrections) in enumerate(summary_df.values):
                            current_row_num = 4 + idx
                            sheet_name = re.sub(r'[\\/?*\[\]:()\'"<>|]', '', anomaly_type).replace(' ', '_').replace('.', '').replace(':', '_').strip(); sheet_name = sheet_name[:31].rstrip('_').strip(); original_sheet_name = sheet_name; s_counter = 1
                            while sheet_name in created_sheet_names: sheet_name = f"{original_sheet_name[:28]}_{s_counter}"; s_counter += 1
                            created_sheet_names.add(sheet_name)
                            summary_cell = ws_summary.cell(row=current_row_num, column=1)
                            summary_cell.hyperlink = f"#'{sheet_name}'!A1"
                            summary_cell.font = Font(underline="single", color="0563C1")
                            ws_detail = wb.create_sheet(title=sheet_name); filtered_df = anomalies_df[anomalies_df['Anomalie'].str.contains(re.escape(anomaly_type), regex=True)]; filtered_df_display = filtered_df.drop(columns=['Anomalie Détaillée FP2E'], errors='ignore')
                            for r in dataframe_to_rows(filtered_df_display, index=False, header=True): ws_detail.append(r)
                            for cell in ws_detail[1]: cell.font = header_font
                            for row_num_detail, df_row_detail in enumerate(filtered_df.iterrows()):
                                for anomaly in str(df_row_detail[1]['Anomalie']).split(' / '):
                                    if anomaly.strip() in anomaly_columns_map:
                                        for col_name in anomaly_columns_map[anomaly.strip()]:
                                            try: ws_detail.cell(row=row_num_detail + 2, column=list(filtered_df_display.columns).index(col_name) + 1).fill = red_fill
                                            except ValueError: pass
                            for col in ws_detail.columns: ws_detail.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(cell.value)) for cell in col if cell.value) + 2
                        wb.save(excel_buffer); st.download_button(label="📥 Télécharger le rapport (.xlsx)", data=excel_buffer, file_name='anomalies_telerelève.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                else: st.success("✅ Aucune anomalie détectée.")
        except Exception as e: st.error(f"Une erreur est survenue : {e}")

# --- ONGLET 3 : CONTROLE MANUELLE ---
with tab3:
    st.header("Contrôle des données manuelles")
    st.markdown("Veuillez téléverser votre fichier pour lancer les contrôles.")
    uploaded_file_manuelle = st.file_uploader("Choisissez un fichier (Manuelle)", type=['csv', 'xlsx'], key="uploader_manuelle")

    if uploaded_file_manuelle:
        st.success("Fichier chargé avec succès !")
        try:
            file_extension = uploaded_file_manuelle.name.split('.')[-1]
            dtype_mapping = {'Numéro de branchement': str, 'Abonnement': str}
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file_manuelle, sep=get_csv_delimiter_radio(uploaded_file_manuelle), dtype=dtype_mapping)
            else:
                df = pd.read_excel(uploaded_file_manuelle, dtype=dtype_mapping)
            
            st.subheader("Aperçu des 5 premières lignes")
            st.dataframe(df.head())

            if st.button("Lancer les contrôles (Manuelle)", key="button_manuelle"):
                with st.spinner("Contrôles en cours..."):
                    anomalies_df, anomaly_counter = check_data_manuelle(df)
                
                if not anomalies_df.empty:
                    st.error(f"Anomalies et/ou corrections détectées : {len(anomalies_df)} lignes concernées.")
                    st.dataframe(anomalies_df)
                    summary_df = create_summary_with_corrections(anomalies_df, anomaly_counter, tab_type="manuelle")
                    st.subheader("Récapitulatif des anomalies"); st.dataframe(summary_df)

                    anomaly_columns_map_manuelle = {
                        "Coordonnées GPS non numériques": ['Latitude', 'Longitude'],
                        "Coordonnées GPS invalides": ['Latitude', 'Longitude'],
                        "L'année de millésime n'est pas conforme": ['Année de fabrication'],
                        "Le diamètre n'est pas conforme": ['Diametre'],
                        "Compteur non-FP2E pour SAPPEL/ITRON": ['Numéro de compteur'],
                        "SAPPEL: Incohérence Marque/Compteur (C)": ['Marque'],
                        "SAPPEL: Incohérence Marque/Compteur (H)": ['Marque'],
                        "ITRON: Incohérence Marque/Compteur": ['Marque']
                    }
                    
                    if file_extension == 'csv':
                        st.download_button(label="📥 Télécharger le rapport en CSV", data=anomalies_df.to_csv(index=False).encode('utf-8'), file_name='anomalies_manuelle.csv', mime='text/csv')
                    elif file_extension == 'xlsx':
                        excel_buffer = io.BytesIO(); wb = Workbook();
                        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
                        ws_summary = wb.create_sheet(title="Récapitulatif", index=0); ws_all_anomalies = wb.create_sheet(title="Toutes_Anomalies", index=1)
                        for r in dataframe_to_rows(anomalies_df, index=False, header=True): ws_all_anomalies.append(r)
                        header_font = Font(bold=True); red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        for cell in ws_all_anomalies[1]: cell.font = header_font
                        for row_num_all, df_row in enumerate(anomalies_df.iterrows()):
                            for anomaly in str(df_row[1]['Anomalie']).split(' / '):
                                if anomaly.strip() in anomaly_columns_map_manuelle:
                                    for col_name in anomaly_columns_map_manuelle[anomaly.strip()]:
                                        try: ws_all_anomalies.cell(row=row_num_all + 2, column=list(anomalies_df.columns).index(col_name) + 1).fill = red_fill
                                        except ValueError: pass
                        for col in ws_all_anomalies.columns: ws_all_anomalies.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(cell.value)) for cell in col if cell.value) + 2
                        ws_summary['A1'] = "Récapitulatif des anomalies"; ws_summary['A1'].font = Font(bold=True, size=16); ws_summary.append([]); 
                        for r in dataframe_to_rows(summary_df, index=False, header=True): ws_summary.append(r)
                        for cell in ws_summary[3]: cell.font = header_font
                        created_sheet_names = {"Récapitulatif", "Toutes_Anomalies"}
                        link_row = ws_summary.max_row + 2; ws_summary.cell(row=link_row, column=1, value="Toutes les anomalies").hyperlink = f"#'Toutes_Anomalies'!A1"; ws_summary.cell(row=link_row, column=1).font = Font(underline="single", color="0563C1"); ws_summary.cell(row=link_row, column=2, value=len(anomalies_df))
                        for idx, (anomaly_type, count, corrections) in enumerate(summary_df.values):
                            current_row_num = 4 + idx
                            sheet_name = re.sub(r'[\\/?*\[\]:()\'"<>|]', '', anomaly_type).replace(' ', '_').replace('.', '').replace(':', '_').strip(); sheet_name = sheet_name[:31].rstrip('_').strip(); original_sheet_name = sheet_name; s_counter = 1
                            while sheet_name in created_sheet_names: sheet_name = f"{original_sheet_name[:28]}_{s_counter}"; s_counter += 1
                            created_sheet_names.add(sheet_name)
                            summary_cell = ws_summary.cell(row=current_row_num, column=1)
                            summary_cell.hyperlink = f"#'{sheet_name}'!A1"
                            summary_cell.font = Font(underline="single", color="0563C1")
                            ws_detail = wb.create_sheet(title=sheet_name); filtered_df = anomalies_df[anomalies_df['Anomalie'].str.contains(re.escape(anomaly_type), regex=True)]
                            for r in dataframe_to_rows(filtered_df, index=False, header=True): ws_detail.append(r)
                            for cell in ws_detail[1]: cell.font = header_font
                            for row_num_detail, df_row_detail in enumerate(filtered_df.iterrows()):
                                for anomaly in str(df_row_detail[1]['Anomalie']).split(' / '):
                                    if anomaly.strip() in anomaly_columns_map_manuelle:
                                        for col_name in anomaly_columns_map_manuelle[anomaly.strip()]:
                                            try: ws_detail.cell(row=row_num_detail + 2, column=list(filtered_df.columns).index(col_name) + 1).fill = red_fill
                                            except ValueError: pass
                            for col in ws_detail.columns: ws_detail.column_dimensions[get_column_letter(col[0].column)].width = max(len(str(cell.value)) for cell in col if cell.value) + 2
                        wb.save(excel_buffer); st.download_button(label="📥 Télécharger le rapport (.xlsx)", data=excel_buffer, file_name='anomalies_manuelle.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                else:
                    st.success("✅ Aucune anomalie détectée.")

        except Exception as e:
            st.error(f"Une erreur est survenue lors du traitement du fichier : {e}")
